#!/usr/bin/env python3
"""Generate SAAP 2026 v3 Excel — final structure.

v3 changes:
- 2 objectives (was 3): Events (80%) + AI Training (20%)
- International expansion folded into both objectives (Option B)
- Revenue target: RM1,000,000 (Events RM800K + AI Training RM200K)
- Outcome-focused KRs throughout
- Best of 32-item (old spreadsheet) + 24-item (current app) versions
- HRDCorp certification, 3 module curriculum, 20 session target from 24-item
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

# ── Color palette ──────────────────────────────────────────────
TEAL = "00897B"
TEAL_LIGHT = "E0F2F1"
DARK = "263238"
WHITE = "FFFFFF"
GRAY_BG = "F5F5F5"
GRAY_BORDER = "CFD8DC"
GREEN = "43A047"
GREEN_LIGHT = "E8F5E9"
AMBER = "FB8C00"
AMBER_LIGHT = "FFF8E1"
RED = "E53935"
RED_LIGHT = "FFEBEE"
BLUE = "1E88E5"
BLUE_LIGHT = "E3F2FD"

header_font = Font(name="Aptos", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
body_font = Font(name="Aptos", color=DARK, size=10)
bold_font = Font(name="Aptos", bold=True, color=DARK, size=10)
thin_border = Border(
    left=Side(style="thin", color=GRAY_BORDER),
    right=Side(style="thin", color=GRAY_BORDER),
    top=Side(style="thin", color=GRAY_BORDER),
    bottom=Side(style="thin", color=GRAY_BORDER),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_body_cell(cell, font=None, fill=None, align=None, fmt=None):
    cell.font = font or body_font
    cell.border = thin_border
    cell.alignment = align or wrap_align
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt


# ═══════════════════════════════════════════════════════════════
# SHEET 1: OKR SUMMARY
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "OKR Summary"

# Title block
ws1.merge_cells("A1:K1")
ws1["A1"].value = "MotionVii SAAP 2026 — OKR Summary"
ws1["A1"].font = Font(name="Aptos", bold=True, color=TEAL, size=16)
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[1].height = 35

ws1.merge_cells("A2:K2")
ws1["A2"].value = "Strategic Annual Action Plan — Scale Events Business & Launch AI Training Revenue Stream"
ws1["A2"].font = Font(name="Aptos", color="78909C", size=11)
ws1["A2"].alignment = Alignment(horizontal="left")

# Revenue target row
ws1.merge_cells("A3:B3")
ws1["A3"].value = "Revenue Target:"
ws1["A3"].font = Font(name="Aptos", bold=True, color=DARK, size=11)
ws1["C3"].value = "RM1,000,000"
ws1["C3"].font = Font(name="Aptos", bold=True, color=TEAL, size=11)
ws1.merge_cells("D3:F3")
ws1["D3"].value = "Events: RM800,000 (80%)  |  AI Training: RM200,000 (20%)"
ws1["D3"].font = Font(name="Aptos", color=DARK, size=10)
ws1.row_dimensions[3].height = 22

# Headers
headers1 = [
    "Obj #", "Objective", "KR #", "Key Result",
    "Target", "Actual", "Unit", "Deadline",
    "Progress %", "Status", "Owner"
]
row = 5
for col, h in enumerate(headers1, 1):
    ws1.cell(row=row, column=col, value=h)
style_header_row(ws1, row, len(headers1))

# ── OKR Data ──────────────────────────────────────────────────
objectives = [
    {
        "num": 1,
        "name": "Scale Events Business",
        "krs": [
            {
                "id": "KR1.1",
                "desc": "Win 6 event contracts generating RM800K+ combined revenue by Q4 2026",
                "target": 800000, "actual": 0, "unit": "RM",
                "deadline": "Q4 2026", "status": "Not Started", "owner": "Khairul"
            },
            {
                "id": "KR1.2",
                "desc": "Establish 3 active event partnerships each producing at least 1 joint proposal by Q3 2026",
                "target": 3, "actual": 0, "unit": "active partnerships",
                "deadline": "Q3 2026", "status": "Not Started", "owner": "Azlan"
            },
            {
                "id": "KR1.3",
                "desc": "Secure 3 repeat bookings or referrals from existing event clients by Q4 2026",
                "target": 3, "actual": 0, "unit": "repeat/referred clients",
                "deadline": "Q4 2026", "status": "Not Started", "owner": "Khairul"
            },
        ]
    },
    {
        "num": 2,
        "name": "Build AI Training Business",
        "krs": [
            {
                "id": "KR2.1",
                "desc": "Deliver 20 paid corporate AI training sessions by Q4 2026",
                "target": 20, "actual": 0, "unit": "paid sessions",
                "deadline": "Q4 2026", "status": "Not Started", "owner": "Khairul"
            },
            {
                "id": "KR2.2",
                "desc": "Generate RM200K in AI training revenue by Q4 2026",
                "target": 200000, "actual": 0, "unit": "RM",
                "deadline": "Q4 2026", "status": "Not Started", "owner": "Khairul"
            },
            {
                "id": "KR2.3",
                "desc": "Secure 5 repeat or referred training clients by Q4 2026",
                "target": 5, "actual": 0, "unit": "repeat/referred clients",
                "deadline": "Q4 2026", "status": "Not Started", "owner": "Khairul"
            },
        ]
    },
]

row = 6
for obj in objectives:
    first_kr_row = row
    for i, kr in enumerate(obj["krs"]):
        progress = round((kr["actual"] / kr["target"]) * 100, 1) if kr["target"] else 0

        c_obj_num = ws1.cell(row=row, column=1, value=obj["num"])
        c_obj_name = ws1.cell(row=row, column=2, value=obj["name"])
        c_kr_id = ws1.cell(row=row, column=3, value=kr["id"])
        c_kr_desc = ws1.cell(row=row, column=4, value=kr["desc"])
        if kr["unit"] == "RM":
            c_target = ws1.cell(row=row, column=5, value=kr["target"])
            c_target.number_format = '#,##0'
            c_actual = ws1.cell(row=row, column=6, value=kr["actual"])
            c_actual.number_format = '#,##0'
        else:
            c_target = ws1.cell(row=row, column=5, value=kr["target"])
            c_actual = ws1.cell(row=row, column=6, value=kr["actual"])
        c_unit = ws1.cell(row=row, column=7, value=kr["unit"])
        c_deadline = ws1.cell(row=row, column=8, value=kr["deadline"])
        c_progress = ws1.cell(row=row, column=9, value=f"{progress}%")
        c_status = ws1.cell(row=row, column=10, value=kr["status"])
        c_owner = ws1.cell(row=row, column=11, value=kr["owner"])

        row_fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid") if row % 2 == 0 else None
        for c in [c_obj_num, c_obj_name, c_kr_id, c_kr_desc, c_target, c_actual,
                   c_unit, c_deadline, c_progress, c_status, c_owner]:
            style_body_cell(c, fill=row_fill)
        style_body_cell(c_obj_name, font=bold_font, fill=row_fill)
        style_body_cell(c_obj_num, font=bold_font, fill=row_fill, align=center_align)
        style_body_cell(c_kr_id, font=bold_font, fill=row_fill, align=center_align)
        for c in [c_target, c_actual, c_unit, c_deadline, c_progress, c_status, c_owner]:
            c.alignment = center_align

        # Status color
        status_colors = {
            "On Track": (GREEN_LIGHT, GREEN),
            "At Risk": (AMBER_LIGHT, AMBER),
            "Behind": (RED_LIGHT, RED),
        }
        if kr["status"] in status_colors:
            bg, fg = status_colors[kr["status"]]
            c_status.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c_status.font = Font(name="Aptos", color=fg, size=10, bold=True)

        row += 1

    if len(obj["krs"]) > 1:
        ws1.merge_cells(start_row=first_kr_row, start_column=1, end_row=row - 1, end_column=1)
        ws1.merge_cells(start_row=first_kr_row, start_column=2, end_row=row - 1, end_column=2)
        ws1.cell(row=first_kr_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=first_kr_row, column=2).alignment = Alignment(vertical="center", wrap_text=True)

ws1.column_dimensions["A"].width = 7
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 8
ws1.column_dimensions["D"].width = 58
ws1.column_dimensions["E"].width = 12
ws1.column_dimensions["F"].width = 10
ws1.column_dimensions["G"].width = 20
ws1.column_dimensions["H"].width = 12
ws1.column_dimensions["I"].width = 12
ws1.column_dimensions["J"].width = 14
ws1.column_dimensions["K"].width = 12
ws1.freeze_panes = "A6"


# ═══════════════════════════════════════════════════════════════
# SHEET 2: KEY RESULTS (detailed tracking)
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Key Results")

ws2.merge_cells("A1:M1")
ws2["A1"].value = "Key Results — Detailed Tracking"
ws2["A1"].font = Font(name="Aptos", bold=True, color=TEAL, size=14)
ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 30

headers2 = [
    "KR ID", "Objective", "Key Result Description",
    "Metric Type", "Target", "Actual", "Unit", "Progress %",
    "Deadline", "Status", "Owner",
    "How We Measure", "Notes"
]
r = 3
for col, h in enumerate(headers2, 1):
    ws2.cell(row=r, column=col, value=h)
style_header_row(ws2, r, len(headers2))

kr_details = [
    # ── Obj 1: Scale Events (80% / RM800K) ────────────────────
    ["KR1.1", "Scale Events", "Win 6 event contracts generating RM800K+ combined revenue by Q4 2026",
     "Revenue", 800000, 0, "RM", "0%", "Q4 2026", "Not Started", "Khairul",
     "Combined signed event contract value in CRM (stage = WON). ~6 contracts averaging RM130K. Includes domestic and international events.",
     "The old KR was 'submit proposals' (activity). Revenue from signed contracts is the outcome. International event wins also count here."],

    ["KR1.2", "Scale Events", "Establish 3 active event partnerships each producing at least 1 joint proposal by Q3 2026",
     "Count", 3, 0, "active partnerships", "0%", "Q3 2026", "Not Started", "Azlan",
     "Signed MOU/agreement AND at least 1 joint proposal or referral generated. 'Active' = commercially productive, not just signed.",
     "Partnerships multiply reach. International partnerships (e.g. through ADIPEC contacts) count."],

    ["KR1.3", "Scale Events", "Secure 3 repeat bookings or referrals from existing event clients by Q4 2026",
     "Count", 3, 0, "repeat/referred clients", "0%", "Q4 2026", "Not Started", "Khairul",
     "Clients who (a) rebook another event or (b) refer a new client. Tracked in CRM with source attribution.",
     "Retention proves quality. Includes year-end appreciation effort to secure 2027 rebookings."],

    # ── Obj 2: AI Training (20% / RM200K) ─────────────────────
    ["KR2.1", "AI Training", "Deliver 20 paid corporate AI training sessions by Q4 2026",
     "Count", 20, 0, "paid sessions", "0%", "Q4 2026", "Not Started", "Khairul",
     "Completed paid sessions (half-day or full-day). Free/discounted pilot workshops do NOT count. HRDCorp-claimable sessions count. International clients count.",
     "20 sessions at ~RM10K avg = RM200K. Pilots (3 planned) are initiatives that validate content, not KR targets."],

    ["KR2.2", "AI Training", "Generate RM200K in AI training revenue by Q4 2026",
     "Revenue", 200000, 0, "RM", "0%", "Q4 2026", "Not Started", "Khairul",
     "Total invoiced AI training revenue. HRDCorp-claimable sessions are premium-priced. 20 sessions x RM10K avg = RM200K target.",
     "HRDCorp certification is a critical initiative — corporates can claim costs from their levy, justifying higher pricing."],

    ["KR2.3", "AI Training", "Secure 5 repeat or referred training clients by Q4 2026",
     "Count", 5, 0, "repeat/referred clients", "0%", "Q4 2026", "Not Started", "Khairul",
     "Clients who either (a) book a second training session, or (b) were referred by a previous client. Tracked in CRM with source.",
     "Repeat/referral = proof of quality. More sustainable than cold acquisition. NPS tracking feeds into this."],
]

r = 4
for kr_row in kr_details:
    for col, val in enumerate(kr_row, 1):
        cell = ws2.cell(row=r, column=col, value=val)
        row_fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid") if r % 2 == 0 else None
        style_body_cell(cell, fill=row_fill)
        if col in (1, 4, 5, 6, 7, 8, 9, 10, 11):
            cell.alignment = center_align
        if col == 1:
            cell.font = bold_font
        if col == 5 and isinstance(val, (int, float)) and val >= 1000:
            cell.number_format = '#,##0'
    r += 1

ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 55
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 10
ws2.column_dimensions["G"].width = 20
ws2.column_dimensions["H"].width = 12
ws2.column_dimensions["I"].width = 12
ws2.column_dimensions["J"].width = 14
ws2.column_dimensions["K"].width = 12
ws2.column_dimensions["L"].width = 60
ws2.column_dimensions["M"].width = 55
ws2.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════
# SHEET 3: INITIATIVES
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Initiatives")

ws3.merge_cells("A1:N1")
ws3["A1"].value = "Initiatives — Action Items"
ws3["A1"].font = Font(name="Aptos", bold=True, color=TEAL, size=14)
ws3["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[1].height = 30

headers3 = [
    "ID", "KR", "Objective", "Initiative",
    "Department", "Start Date", "End Date",
    "Budget (RM)", "Resources", "Person In Charge", "Accountable",
    "Status", "Progress", "Remarks"
]
r = 3
for col, h in enumerate(headers3, 1):
    ws3.cell(row=r, column=col, value=h)
style_header_row(ws3, r, len(headers3))

initiatives = [
    # ══════════════════════════════════════════════════════════
    # OBJECTIVE 1: SCALE EVENTS BUSINESS (80% / RM800K)
    # ══════════════════════════════════════════════════════════

    # ── KR1.1: Win 6 event contracts / RM800K revenue ────────

    [1, "KR1.1", "Scale Events",
     "Identify 15 potential event clients from existing network and industry contacts",
     "Operations", date(2026,1,1), date(2026,2,28),
     None, None, "Azlan", "Khairul", "Pending", "", ""],

    [2, "KR1.1", "Scale Events",
     "Submit 7 event proposals and set client discussions",
     "Operations", date(2026,1,1), date(2026,6,30),
     1400, "Meetings", "Azlan", "Khairul", "Pending", "",
     "Was KR1.1 in old SAAP — now an initiative. Submitting is the work; winning is the outcome."],

    [3, "KR1.1", "Scale Events",
     "Follow-up with potential clients (structured cadence via CRM)",
     "Operations", date(2026,1,1), date(2026,9,30),
     None, None, "Khairul", "Khairul", "Pending", "", ""],

    [4, "KR1.1", "Scale Events",
     "Implement CRM pipeline to track event proposals, follow-ups, and win/loss",
     "Operations", date(2026,1,1), date(2026,3,31),
     1800, "Software - Pipeline", "Khairul", "Khairul", "Pending", "",
     "SAAP can feed data back — track proposal stage, conversion rate"],

    [5, "KR1.1", "Scale Events",
     "Conduct discussion on SUKMA 2026 opportunities with Bisabi Sdn Bhd",
     "Business Dev", date(2026,1,1), date(2026,2,28),
     None, None, "Khairul", "Khairul", "Pending", "", ""],

    [6, "KR1.1", "Scale Events",
     "Attend 1 international exhibition (ADIPEC/OTC Asia) to source international event leads",
     "Business Dev", date(2026,1,1), date(2026,9,30),
     10000, "Travel + registration", "Khairul", "Khairul", "Pending", "",
     "International folded into events. O&G exhibitions have highest-value leads."],

    [7, "KR1.1", "Scale Events",
     "Follow-up with ADIPEC leads and PETRONAS Abu Dhabi on content localisation",
     "Business Dev", date(2026,1,15), date(2026,3,31),
     None, None, "Khairul", "Khairul", "Pending", "",
     "Warmest international lead — Petronas connection in UAE"],

    [8, "KR1.1", "Scale Events",
     "Offer pilot engagement or proof-of-concept to 3 international prospects",
     "Operations", date(2026,2,1), date(2026,6,30),
     None, None, "Khairul", "Khairul", "Pending", "",
     "Low-commitment entry point to demonstrate quality to international market"],

    # ── KR1.2: 3 active event partnerships ───────────────────

    [9, "KR1.2", "Scale Events",
     "Source and approach 5 event organisers (not logistics)",
     "Business Dev", date(2026,1,1), date(2026,3,31),
     None, None, "Azlan", "Khairul", "Pending", "",
     "Need 5 approaches to land 3 active partnerships"],

    [10, "KR1.2", "Scale Events",
     "Set meetings with top 3 event organisers from sourcing list",
     "Business Dev", date(2026,2,1), date(2026,4,30),
     None, None, "Azlan", "Khairul", "Pending", "", ""],

    [11, "KR1.2", "Scale Events",
     "Obtain pricing from 2 event logistics companies for partnership bundling",
     "Business Dev", date(2026,4,1), date(2026,6,30),
     None, None, "Azlan", "Khairul", "Pending", "", ""],

    [12, "KR1.2", "Scale Events",
     "Secure 2 reliable event PIC/Account managers (freelance/contract)",
     "Business Dev", date(2026,1,1), date(2026,6,30),
     60000, "Event manager contracts", "Khairul", "Khairul", "Pending", "", ""],

    [13, "KR1.2", "Scale Events",
     "Attend 2 industry events for networking and partnership development",
     "Business Dev", date(2026,3,1), date(2026,9,30),
     2000, "Delegate passes", "Khairul", "Khairul", "Pending", "", ""],

    [14, "KR1.2", "Scale Events",
     "Create international marketing pack showcasing Petronas and O&G video portfolio",
     "Marketing", date(2026,1,15), date(2026,3,31),
     None, None, "Azlan", "Khairul", "Pending", "",
     "Leverage strongest asset to open both domestic and international partnership doors"],

    # ── KR1.3: 3 repeat bookings / referrals ─────────────────

    [15, "KR1.3", "Scale Events",
     "Build event case studies for marketing (min 2 case studies)",
     "Marketing", date(2026,3,1), date(2026,6,30),
     None, None, "Azlan", "Khairul", "Pending", "",
     "Case studies drive both repeat bookings and new referrals"],

    [16, "KR1.3", "Scale Events",
     "Secure repeat bookings from existing clients through proactive outreach",
     "Business Dev", date(2026,6,1), date(2026,10,31),
     None, None, "Khairul", "Khairul", "Pending", "", ""],

    [17, "KR1.3", "Scale Events",
     "Year-end client appreciation and 2027 planning sessions",
     "Business Dev", date(2026,10,1), date(2026,12,31),
     None, None, "Khairul", "Khairul", "Pending", "",
     "Lock in 2027 commitments while relationship is warm"],

    [18, "KR1.3", "Scale Events",
     "Build 1 international portfolio piece leveraging Petronas partnership",
     "Business Dev", date(2026,3,1), date(2026,9,30),
     None, None, "Azlan", "Khairul", "Pending", "",
     "International credibility piece that feeds both new leads and repeat trust"],

    # ══════════════════════════════════════════════════════════
    # OBJECTIVE 2: BUILD AI TRAINING BUSINESS (20% / RM200K)
    # ══════════════════════════════════════════════════════════

    # ── KR2.1: Deliver 20 paid corporate sessions ────────────

    [19, "KR2.1", "AI Training",
     "Research corporate AI training market, competitors, and pricing in Malaysia",
     "Business Dev", date(2026,1,1), date(2026,2,15),
     None, None, "Azlan", "Khairul", "Pending", "",
     "What would Petronas/corporate teams actually pay to learn?"],

    [20, "KR2.1", "AI Training",
     "Define 3 AI training modules (topics, duration, pricing)",
     "Business Dev", date(2026,1,15), date(2026,2,28),
     None, None, "Khairul", "Khairul", "Pending", "",
     "e.g. AI for O&G, AI for Marketing, Generative AI Workflows"],

    [21, "KR2.1", "AI Training",
     "Develop Module 1 curriculum and materials",
     "Operations", date(2026,2,1), date(2026,3,31),
     5000, "Content development", "Khairul", "Khairul", "Pending", "", ""],

    [22, "KR2.1", "AI Training",
     "Develop Module 2 curriculum and materials",
     "Operations", date(2026,3,1), date(2026,4,30),
     5000, "Content development", "Khairul", "Khairul", "Pending", "", ""],

    [23, "KR2.1", "AI Training",
     "Develop Module 3 curriculum and materials",
     "Operations", date(2026,4,1), date(2026,5,31),
     5000, "Content development", "Khairul", "Khairul", "Pending", "", ""],

    [24, "KR2.1", "AI Training",
     "Conduct 3 pilot workshops (discounted/free) to validate content and delivery",
     "Operations", date(2026,3,1), date(2026,5,31),
     None, None, "Khairul", "Khairul", "Pending", "",
     "Pilots do NOT count toward KR2.1 (paid sessions). They validate content and generate testimonials."],

    [25, "KR2.1", "AI Training",
     "Refine curriculum based on pilot feedback",
     "Operations", date(2026,5,1), date(2026,6,30),
     None, None, "Khairul", "Khairul", "Pending", "",
     "Iterate before commercial launch — critical quality gate"],

    # ── KR2.2: Generate RM200K in AI training revenue ────────

    [26, "KR2.2", "AI Training",
     "Apply for HRDCorp training provider certification",
     "Business Dev", date(2026,1,15), date(2026,4,30),
     None, None, "Khairul", "Khairul", "Pending", "",
     "CRITICAL: HRDCorp certification lets corporates claim training costs from levy. Major selling point in Malaysia."],

    [27, "KR2.2", "AI Training",
     "Price and package training offerings (half-day, full-day, multi-session series)",
     "Operations", date(2026,2,1), date(2026,3,31),
     None, None, "Khairul", "Khairul", "Pending", "",
     "Target RM10K+ per corporate session. HRDCorp-claimable sessions priced at premium."],

    [28, "KR2.2", "AI Training",
     "Identify 20 target companies for AI training (leverage Petronas network)",
     "Business Dev", date(2026,2,1), date(2026,4,30),
     None, None, "Azlan", "Khairul", "Pending", "",
     "Start with warm leads — companies that already know MotionVii or Petronas ecosystem"],

    [29, "KR2.2", "AI Training",
     "Create AI Training marketing collateral (brochure, website section)",
     "Marketing", date(2026,3,1), date(2026,5,31),
     3000, "Design + web", "Azlan", "Khairul", "Pending", "", ""],

    [30, "KR2.2", "AI Training",
     "Launch full commercial AI training offering",
     "Marketing", date(2026,5,1), date(2026,6,30),
     None, None, "Khairul", "Khairul", "Pending", "",
     "After pilots are validated and HRDCorp cert is in progress"],

    [31, "KR2.2", "AI Training",
     "Pitch AI training to top 10 corporate prospects",
     "Business Dev", date(2026,4,1), date(2026,7,31),
     None, None, "Khairul", "Khairul", "Pending", "", ""],

    [32, "KR2.2", "AI Training",
     "Deliver 20 corporate workshop sessions",
     "Operations", date(2026,5,1), date(2026,12,31),
     None, None, "Khairul", "Khairul", "Pending", "",
     "Main delivery effort — ramp up after commercial launch"],

    [33, "KR2.2", "AI Training",
     "Explore AI training opportunities with international corporates (e.g. Petronas Abu Dhabi)",
     "Business Dev", date(2026,6,1), date(2026,12,31),
     None, None, "Khairul", "Khairul", "Pending", "",
     "International folded in — international training revenue also counts toward 20%"],

    # ── KR2.3: 5 repeat or referred training clients ─────────

    [34, "KR2.3", "AI Training",
     "Implement post-training feedback survey and NPS tracking",
     "Operations", date(2026,4,1), date(2026,5,31),
     None, None, "Azlan", "Khairul", "Pending", "",
     "Data to prove quality, identify improvements, and support testimonials"],

    [35, "KR2.3", "AI Training",
     "Create AI training case study and testimonials from first sessions",
     "Marketing", date(2026,6,1), date(2026,8,31),
     None, None, "Azlan", "Khairul", "Pending", "",
     "Social proof drives referrals — get permission from pilot participants"],

    [36, "KR2.3", "AI Training",
     "Follow up with all trained participants about advanced sessions and referrals",
     "Business Dev", date(2026,7,1), date(2026,12,31),
     None, None, "Khairul", "Khairul", "Pending", "",
     "Systematic — don't rely on organic referrals. Schedule follow-ups in CRM."],

    [37, "KR2.3", "AI Training",
     "Offer referral incentive (discount on next session) to existing training clients",
     "Marketing", date(2026,6,1), date(2026,12,31),
     None, None, "Azlan", "Khairul", "Pending", "", ""],
]

r = 4
for init in initiatives:
    for col, val in enumerate(init, 1):
        cell = ws3.cell(row=r, column=col, value=val)
        row_fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid") if r % 2 == 0 else None
        style_body_cell(cell, fill=row_fill)

        if col == 1:
            cell.alignment = center_align
            cell.font = bold_font
        elif col == 2:
            cell.alignment = center_align
            cell.font = bold_font
        elif col in (5, 10, 11, 12, 13):
            cell.alignment = center_align
        elif col == 8:
            if val:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="top")
        elif col in (6, 7):
            if val:
                cell.number_format = 'DD MMM YYYY'
            cell.alignment = center_align
    r += 1

ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 8
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 62
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 14
ws3.column_dimensions["G"].width = 14
ws3.column_dimensions["H"].width = 13
ws3.column_dimensions["I"].width = 22
ws3.column_dimensions["J"].width = 16
ws3.column_dimensions["K"].width = 14
ws3.column_dimensions["L"].width = 12
ws3.column_dimensions["M"].width = 10
ws3.column_dimensions["N"].width = 55
ws3.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════
# SHEET 4: STRUCTURE GUIDE
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Structure Guide")

ws4.merge_cells("A1:F1")
ws4["A1"].value = "OKR Structure Guide"
ws4["A1"].font = Font(name="Aptos", bold=True, color=TEAL, size=14)
ws4.row_dimensions[1].height = 30

guide_content = [
    ["", "", "", "", "", ""],
    ["Layer", "What It Is", "Example", "Tracking", "Review Cadence", "Who Owns It"],
    [
        "Objective",
        "Aspirational direction\n(qualitative, inspiring)\n\n2 objectives for 2026:\n80% Events + 20% AI Training",
        "Scale Events Business",
        "No metric — just direction.\nSuccess = KRs achieved.",
        "Annual",
        "CEO / Leadership"
    ],
    [
        "Key Result",
        "Measurable OUTCOME\n(specific, time-bound)\n\nMust answer: 'What changed?'\nNOT: 'What did we do?'",
        "Win 6 event contracts\ngenerating RM800K+ by Q4",
        "Target vs Actual\nProgress %\nStatus (On Track / At Risk)",
        "Monthly review\nQuarterly scoring",
        "KR Owner"
    ],
    [
        "Initiative",
        "Action item that DRIVES a KR\n\nThis is the work you do.\nMultiple initiatives per KR.",
        "Submit 7 event proposals\nand set discussions",
        "Status (Pending → Completed)\nStart/End dates, Budget",
        "Weekly / Bi-weekly",
        "Person In Charge"
    ],
    ["", "", "", "", "", ""],
    ["Revenue Target Breakdown:", "", "", "", "", ""],
    ["Stream", "Target", "% of Total", "Key Revenue Logic", "", ""],
    [
        "Events",
        "RM800,000",
        "80%",
        "~6 events x RM130K avg.\nIncludes domestic + international.",
        "", ""
    ],
    [
        "AI Training",
        "RM200,000",
        "20%",
        "~20 sessions x RM10K avg.\nHRDCorp-claimable = premium pricing.",
        "", ""
    ],
    [
        "Total",
        "RM1,000,000",
        "100%",
        "International revenue counts\nunder whichever stream it falls.",
        "", ""
    ],
    ["", "", "", "", "", ""],
    ["What Changed from Previous SAAP:", "", "", "", "", ""],
    ["Change", "Before", "After", "Why", "", ""],
    [
        "3 → 2 Objectives",
        "Events + International B2B\n+ AI Product (3 separate)",
        "Events (80%) + AI Training (20%)\nInternational folded into both",
        "International is a market strategy,\nnot a business line. Revenue counts\nunder events or training.",
        "", ""
    ],
    [
        "AI Product → Training",
        "Build product, join accelerator,\nfind investor",
        "Deliver paid training to corporates,\ngenerate RM200K, earn repeat clients",
        "Training leverages existing expertise\nand client base (Petronas network).\nNo investor needed — revenue model.",
        "", ""
    ],
    [
        "KR = Outcome",
        "'Submit 7 proposals' (activity)\n'Build lead system' (task)",
        "'Win 6 contracts / RM800K' (outcome)\n'Deliver 20 sessions' (outcome)",
        "KRs measure results, not effort.\nActivities become initiatives.",
        "", ""
    ],
    [
        "HRDCorp certification",
        "Not in previous SAAP",
        "Critical initiative under KR2.2\n(enables corporate sales)",
        "Corporates claim training costs\nfrom HRDCorp levy — major\nselling point in Malaysian market.",
        "", ""
    ],
    [
        "Revenue targets",
        "No explicit revenue numbers\non KRs",
        "KR1.1: RM800K | KR2.2: RM200K\nRM1M total",
        "Revenue makes KRs concrete\nand ties directly to business goal.",
        "", ""
    ],
    ["", "", "", "", "", ""],
    ["OKR Anti-Patterns to Avoid:", "", "", "", "", ""],
    ["Anti-Pattern", "Example", "Fix", "Diagnostic Test", "", ""],
    [
        "Activity as KR",
        "'Submit 7 proposals'",
        "'Win 4 contracts'",
        "'Can I do this without the\nbusiness improving?'\nIf yes → it's an activity.",
        "", ""
    ],
    [
        "Task as KR",
        "'Build lead capture system'",
        "'Generate 20 qualified leads'",
        "'Is this a deliverable or a result?'\nSystems/tools are deliverables.",
        "", ""
    ],
    [
        "Vague KR",
        "'Develop 3 partnerships'",
        "'3 active partnerships each\nproducing 1+ joint proposal'",
        "'Would two people agree this\nis achieved?' Add qualifying criteria.",
        "", ""
    ],
]

for ridx, row_data in enumerate(guide_content, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=ridx, column=c, value=val)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.border = thin_border

        # Header rows (Layer/Revenue/Changes/Anti-pattern tables)
        if ridx in (3, 9, 15, 23):
            cell.font = header_font
            cell.fill = header_fill
        # Section title rows
        elif ridx in (8, 14, 22):
            cell.font = Font(name="Aptos", bold=True, color=TEAL, size=11)
            cell.border = Border()
        # "Total" row bold
        elif ridx == 12:
            cell.font = bold_font

ws4.column_dimensions["A"].width = 24
ws4.column_dimensions["B"].width = 30
ws4.column_dimensions["C"].width = 30
ws4.column_dimensions["D"].width = 38
ws4.column_dimensions["E"].width = 14
ws4.column_dimensions["F"].width = 14


# ═══════════════════════════════════════════════════════════════
# SHEET 5: SUPPORT TASKS
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Support Tasks")

ws5.merge_cells("A1:H1")
ws5["A1"].value = "Support Tasks — Operational Work Supporting SAAP Initiatives"
ws5["A1"].font = Font(name="Aptos", bold=True, color=TEAL, size=14)
ws5["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws5.row_dimensions[1].height = 30

ws5.merge_cells("A2:H2")
ws5["A2"].value = "These are recurring, ad-hoc, or BAU tasks — not strategic initiatives, but needed to deliver them."
ws5["A2"].font = Font(name="Aptos", color="78909C", size=10)
ws5["A2"].alignment = Alignment(horizontal="left")

headers5 = [
    "ID", "Category", "Task", "Supports",
    "Owner", "Frequency", "Priority", "Notes"
]
r = 4
for col, h in enumerate(headers5, 1):
    ws5.cell(row=r, column=col, value=h)
style_header_row(ws5, r, len(headers5))

# Category colors
CAT_DESIGN = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")  # Indigo light
CAT_BIZ = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")     # Teal light
CAT_TALENTA = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # Orange light
CAT_OPS = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")      # Purple light

support_tasks = [
    # ── DESIGN & CREATIVE (Azlan) ─────────────────────────────
    # category_fill, id, category, task, supports, owner, frequency, priority, notes

    (CAT_DESIGN, 1, "Design & Creative", "Design event proposal decks and pitch materials",
     "KR1.1", "Azlan", "Per proposal", "High",
     "Each of the 7 proposals needs a tailored deck. Template-first approach to save time."),

    (CAT_DESIGN, 2, "Design & Creative", "Design event collateral (banners, backdrops, signage, badges)",
     "KR1.1", "Azlan", "Per event won", "High",
     "Triggered after contract signed. Budget included in event project cost."),

    (CAT_DESIGN, 3, "Design & Creative", "Video production and editing for event content",
     "KR1.1, KR1.3", "Azlan", "Per event", "High",
     "Event recap videos, highlight reels. Also feeds case studies for KR1.3."),

    (CAT_DESIGN, 4, "Design & Creative", "Design event case study layouts",
     "KR1.3", "Azlan", "2x per year", "Medium",
     "Supports initiative #15 — need min 2 case studies for marketing."),

    (CAT_DESIGN, 5, "Design & Creative", "Design international marketing pack (portfolio, showreel)",
     "KR1.2", "Azlan", "Once + updates", "High",
     "Supports initiative #14. Showcase Petronas/O&G work for partnerships and intl prospects."),

    (CAT_DESIGN, 6, "Design & Creative", "Update MotionVii website content and visuals",
     "KR1.2, KR2.2", "Azlan", "Monthly", "Medium",
     "Keep portfolio current. Add AI training section after commercial launch."),

    (CAT_DESIGN, 7, "Design & Creative", "Social media content creation (posts, reels, stories)",
     "KR1.2, KR2.2", "Azlan", "Weekly", "Medium",
     "Maintain brand visibility across both events and AI training."),

    (CAT_DESIGN, 8, "Design & Creative", "Design AI training presentation slides and handout materials",
     "KR2.1", "Azlan", "Per module", "High",
     "Supports Module 1/2/3 development. Professional materials = premium pricing justified."),

    (CAT_DESIGN, 9, "Design & Creative", "Design AI training brochure and marketing collateral",
     "KR2.2", "Azlan", "Once + updates", "High",
     "Supports initiative #29. Needs to be ready before commercial launch."),

    (CAT_DESIGN, 10, "Design & Creative", "Video testimonials from training participants",
     "KR2.3", "Azlan", "After each session", "Medium",
     "Quick video testimonials boost referral credibility. Get permission during session."),

    (CAT_DESIGN, 11, "Design & Creative", "Portfolio and showreel updates (quarterly refresh)",
     "KR1.1, KR1.2", "Azlan", "Quarterly", "Medium",
     "Keep demo reel current with latest event and video work."),

    # ── BUSINESS & ADMIN (Khairul) ────────────────────────────

    (CAT_BIZ, 12, "Business & Admin", "Proposal writing and quotation preparation",
     "KR1.1", "Khairul", "Per opportunity", "High",
     "Each proposal needs custom scope, pricing, timeline. Use templates to speed up."),

    (CAT_BIZ, 13, "Business & Admin", "Contract preparation, review, and execution",
     "KR1.1, KR2.2", "Khairul", "Per deal", "High",
     "Both event contracts and AI training contracts."),

    (CAT_BIZ, 14, "Business & Admin", "Invoicing and payment follow-up",
     "KR1.1, KR2.2", "Khairul", "Per project/session", "High",
     "Revenue only counts when invoiced. Track in SAAP project financials."),

    (CAT_BIZ, 15, "Business & Admin", "Financial reporting and budget tracking (SAAP)",
     "All KRs", "Khairul", "Monthly", "High",
     "Monthly review of revenue vs RM1M target. Events vs Training split."),

    (CAT_BIZ, 16, "Business & Admin", "Client relationship management and meeting notes",
     "KR1.1, KR1.3", "Khairul", "Ongoing", "Medium",
     "CRM updates after every client interaction. Feed data to KR tracking."),

    (CAT_BIZ, 17, "Business & Admin", "Supplier and vendor management (logistics, venues, AV)",
     "KR1.1, KR1.2", "Khairul", "Per event", "Medium",
     "Negotiate rates, manage relationships, ensure delivery quality."),

    (CAT_BIZ, 18, "Business & Admin", "HR — freelancer/contractor onboarding and management",
     "KR1.2", "Khairul", "As needed", "Medium",
     "Event PIC managers, marketing contractors. Supports initiative #12."),

    (CAT_BIZ, 19, "Business & Admin", "HRDCorp application documentation and follow-up",
     "KR2.2", "Khairul", "Until approved", "High",
     "Supports initiative #26. Documentation-heavy process — track milestones."),

    (CAT_BIZ, 20, "Business & Admin", "AI training session logistics (venue, equipment, catering)",
     "KR2.1", "Khairul", "Per session", "Medium",
     "Book venue, arrange equipment, handle logistics for each corporate session."),

    (CAT_BIZ, 21, "Business & Admin", "Partnership agreement drafting and negotiation",
     "KR1.2", "Khairul", "Per partnership", "Medium",
     "MOU or formal partnership agreement. Legal review if needed."),

    (CAT_BIZ, 22, "Business & Admin", "International compliance and logistics (travel, permits, banking)",
     "KR1.1", "Khairul", "Per intl engagement", "Low",
     "Only triggered when international work materializes. Cross-border invoicing, travel planning."),

    # ── TALENTA IDEAS REQUESTS ────────────────────────────────

    (CAT_TALENTA, 23, "Talenta Ideas", "Design requests from Talenta Ideas (ad-hoc)",
     "Parent company", "Azlan", "Ad-hoc", "Medium",
     "As subsidiary, MotionVii supports Talenta's design needs. Track hours to manage capacity."),

    (CAT_TALENTA, 24, "Talenta Ideas", "Video production requests from Talenta Ideas",
     "Parent company", "Azlan", "Ad-hoc", "Medium",
     "Corporate videos, internal comms, event coverage for Talenta's own projects."),

    (CAT_TALENTA, 25, "Talenta Ideas", "Talenta Ideas brand and marketing material updates",
     "Parent company", "Azlan", "Quarterly", "Low",
     "Brochure updates, presentation templates, brand guideline maintenance."),

    (CAT_TALENTA, 26, "Talenta Ideas", "Coordination and reporting to Talenta Ideas management",
     "Parent company", "Khairul", "Monthly", "Medium",
     "Monthly update on MotionVii performance, revenue, and SAAP progress to parent company."),

    # ── OPERATIONS & INFRASTRUCTURE ───────────────────────────

    (CAT_OPS, 27, "Operations", "SAAP platform maintenance and development",
     "All KRs", "Khairul", "Ongoing", "Medium",
     "The tool tracking all of this. Bug fixes, new features, data integrity."),

    (CAT_OPS, 28, "Operations", "Software subscriptions and tool management",
     "All KRs", "Khairul", "Monthly", "Low",
     "CRM, automation tools, design software, project management. Renewals and cost control."),

    (CAT_OPS, 29, "Operations", "Document management and filing (contracts, proposals, invoices)",
     "All KRs", "Khairul", "Ongoing", "Low",
     "Keep organized for audits, HRDCorp requirements, and client records."),

    (CAT_OPS, 30, "Operations", "Team capacity planning and workload balancing",
     "All KRs", "Khairul", "Bi-weekly", "High",
     "3-person team running 37 initiatives + support tasks + Talenta requests. Watch for bottlenecks on Azlan (design) and Khairul (business)."),
]

r = 5
for task in support_tasks:
    cat_fill = task[0]
    data = task[1:]  # id, category, task, supports, owner, frequency, priority, notes

    for col, val in enumerate(data, 1):
        cell = ws5.cell(row=r, column=col, value=val)
        style_body_cell(cell)

        # Apply category color to row
        cell.fill = cat_fill

        if col == 1:  # ID
            cell.alignment = center_align
            cell.font = bold_font
        elif col == 2:  # Category
            cell.font = bold_font
        elif col in (4, 5, 6, 7):  # Supports, Owner, Frequency, Priority
            cell.alignment = center_align

        # Priority color override
        if col == 7:
            if val == "High":
                cell.font = Font(name="Aptos", color=RED, size=10, bold=True)
            elif val == "Medium":
                cell.font = Font(name="Aptos", color=AMBER, size=10, bold=True)
            elif val == "Low":
                cell.font = Font(name="Aptos", color="78909C", size=10)
    r += 1

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 18
ws5.column_dimensions["C"].width = 55
ws5.column_dimensions["D"].width = 16
ws5.column_dimensions["E"].width = 12
ws5.column_dimensions["F"].width = 16
ws5.column_dimensions["G"].width = 10
ws5.column_dimensions["H"].width = 60
ws5.freeze_panes = "A5"


# ── Save ───────────────────────────────────────────────────────
output_path = "/Users/khairul/Documents/MyDev/Work/Motionvii/SAAP2026v2/MotionVii_SAAP_2026_v2.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
print(f"  Objectives: {len(objectives)}")
print(f"  Key Results: {sum(len(o['krs']) for o in objectives)}")
print(f"  Initiatives: {len(initiatives)}")
print(f"  Support Tasks: {len(support_tasks)}")
